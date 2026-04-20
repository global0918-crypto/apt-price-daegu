"""
7년치 아파트 실거래 데이터 수집 (최초 1회 실행)
결과: data/아파트실거래_마스터.xlsx  ← generate_html.py와 apt_trade_fetcher.py가 공유
"""
import requests
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from dateutil.relativedelta import relativedelta
import os
import time

API_KEY = os.environ.get("API_KEY", "c0fb777201beb56b4c5e333b94b2e0d3771cfd15247049464ed6161eb19ee6b3")
API_URL = "https://apis.data.go.kr/1613000/RTMSDataSvcAptTrade/getRTMSDataSvcAptTrade"
MASTER_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "아파트실거래_마스터.xlsx")

REGIONS = {
    "서울": [
        ("종로구","11110"),("중구","11140"),("용산구","11170"),
        ("성동구","11200"),("광진구","11215"),("동대문구","11230"),
        ("중랑구","11260"),("성북구","11290"),("강북구","11305"),
        ("도봉구","11320"),("노원구","11350"),("은평구","11380"),
        ("서대문구","11410"),("마포구","11440"),("양천구","11470"),
        ("강서구","11500"),("구로구","11530"),("금천구","11545"),
        ("영등포구","11560"),("동작구","11590"),("관악구","11620"),
        ("서초구","11650"),("강남구","11680"),("송파구","11710"),
        ("강동구","11740"),
    ],
    "대구": [
        ("중구","27110"),("동구","27140"),("서구","27170"),
        ("남구","27200"),("북구","27230"),("수성구","27260"),
        ("달서구","27290"),("달성군","27710"),
    ],
}

HEADERS = ["시도","구군","법정동","아파트명","전용면적(㎡)","층","건축년도",
           "계약년","계약월","계약일","거래금액(만원)","거래유형","중개사소재지","수집일시"]


def month_range(years=7):
    now = datetime.now()
    start = now - relativedelta(years=years)
    months = []
    cur = start.replace(day=1)
    while cur <= now:
        months.append(cur.strftime("%Y%m"))
        cur += relativedelta(months=1)
    return months


def fetch_all_pages(lawd_cd, deal_ymd):
    items = []
    page = 1
    while True:
        params = {
            "serviceKey": API_KEY,
            "LAWD_CD": lawd_cd,
            "DEAL_YMD": deal_ymd,
            "pageNo": page,
            "numOfRows": 1000,
        }
        try:
            resp = requests.get(API_URL, params=params, timeout=30)
            resp.raise_for_status()
            root = ET.fromstring(resp.text)
            page_items = root.findall(".//item")
            items.extend(page_items)
            total = int(root.findtext(".//totalCount") or 0)
            if page * 1000 >= total:
                break
            page += 1
            time.sleep(0.1)
        except Exception as e:
            print(f"    [오류] {lawd_cd} {deal_ymd} p{page}: {e}")
            break
    return items


def get_text(item, tag):
    el = item.find(tag)
    return el.text.strip() if el is not None and el.text else ""


def load_existing(path):
    """마스터 파일이 있으면 기존 데이터 로드, 없으면 빈 dict 반환"""
    existing = {"서울": set(), "대구": set()}
    rows = {"서울": [], "대구": []}
    if not os.path.exists(path):
        return existing, rows

    wb = openpyxl.load_workbook(path)
    for city in ["서울", "대구"]:
        if city not in wb.sheetnames:
            continue
        ws = wb[city]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rows[city].append(list(row))
            # 중복 키: 아파트명+계약년+계약월+계약일+거래금액+층
            key = (str(row[3]), str(row[7]), str(row[8]), str(row[9]), str(row[10]), str(row[5]))
            existing[city].add(key)
    print(f"기존 데이터: 서울 {len(rows['서울'])}건 / 대구 {len(rows['대구'])}건")
    return existing, rows


def save_master(rows, path):
    wb = openpyxl.Workbook()
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF", size=10)
    align = Alignment(horizontal="center", vertical="center")

    for city in ["서울", "대구"]:
        ws = wb.create_sheet(city)
        ws.append(HEADERS)
        for row in rows[city]:
            ws.append(row)
        for cell in ws[1]:
            cell.fill = fill; cell.font = font; cell.alignment = align
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(path)


def main():
    os.makedirs(os.path.dirname(MASTER_FILE), exist_ok=True)
    months = month_range(7)
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    print(f"=== 7년치 수집 시작 ({months[0]} ~ {months[-1]}) ===")
    print(f"총 {len(months)}개월 × (서울 25구 + 대구 8구)\n")

    existing, rows = load_existing(MASTER_FILE)
    new_count = {"서울": 0, "대구": 0}

    for city, districts in REGIONS.items():
        print(f"\n[{city}]")
        for district_name, lawd_cd in districts:
            print(f"  {district_name}", end="", flush=True)
            dist_new = 0
            for ym in months:
                items = fetch_all_pages(lawd_cd, ym)
                for item in items:
                    apt   = get_text(item, "aptNm")
                    yr    = get_text(item, "dealYear")
                    mo    = get_text(item, "dealMonth")
                    day   = get_text(item, "dealDay")
                    amt   = get_text(item, "dealAmount").replace(",", "")
                    floor = get_text(item, "floor")
                    key   = (apt, yr, mo, day, amt, floor)
                    if key in existing[city]:
                        continue
                    existing[city].add(key)
                    rows[city].append([
                        city, district_name,
                        get_text(item, "umdNm"), apt,
                        get_text(item, "excluUseAr"), floor,
                        get_text(item, "buildYear"),
                        yr, mo, day, amt,
                        get_text(item, "dealingGbn"),
                        get_text(item, "estateAgentSggNm"),
                        now_str,
                    ])
                    dist_new += 1
                time.sleep(0.15)
            print(f" +{dist_new}건", flush=True)
            new_count[city] += dist_new

            # 구 단위로 중간저장
            for c in ["서울", "대구"]:
                rows[c].sort(key=lambda r: (str(r[7]), str(r[8]).zfill(2), str(r[9]).zfill(2)))
            save_master(rows, MASTER_FILE)
            print(f"  [중간저장] 서울 {len(rows['서울'])}건 / 대구 {len(rows['대구'])}건", flush=True)

    print(f"\n신규 수집: 서울 {new_count['서울']}건 / 대구 {new_count['대구']}건")
    print(f"전체 누적: 서울 {len(rows['서울'])}건 / 대구 {len(rows['대구'])}건")

    # 날짜순 정렬
    for city in ["서울", "대구"]:
        rows[city].sort(key=lambda r: (str(r[7]), str(r[8]).zfill(2), str(r[9]).zfill(2)))

    save_master(rows, MASTER_FILE)
    print(f"\n저장 완료: {MASTER_FILE}")


if __name__ == "__main__":
    main()
