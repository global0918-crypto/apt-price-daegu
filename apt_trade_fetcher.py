import requests
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os

API_KEY = os.environ.get("API_KEY", "cad2afdb0f315b4ef965de57205ab4711e2b0ee7788caf4387c5edcd6820c430")
API_URL = "https://apis.data.go.kr/1613000/RTMSDataSvcAptTrade/getRTMSDataSvcAptTrade"
EMAIL_TO = "global0918@gmail.com"
MASTER_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "아파트실거래_마스터.xlsx")

REGIONS = {
    "서울": [
        ("종로구", "11110"), ("중구", "11140"), ("용산구", "11170"),
        ("성동구", "11200"), ("광진구", "11215"), ("동대문구", "11230"),
        ("중랑구", "11260"), ("성북구", "11290"), ("강북구", "11305"),
        ("도봉구", "11320"), ("노원구", "11350"), ("은평구", "11380"),
        ("서대문구", "11410"), ("마포구", "11440"), ("양천구", "11470"),
        ("강서구", "11500"), ("구로구", "11530"), ("금천구", "11545"),
        ("영등포구", "11560"), ("동작구", "11590"), ("관악구", "11620"),
        ("서초구", "11650"), ("강남구", "11680"), ("송파구", "11710"),
        ("강동구", "11740"),
    ],
    "대구": [
        ("중구", "27110"), ("동구", "27140"), ("서구", "27170"),
        ("남구", "27200"), ("북구", "27230"), ("수성구", "27260"),
        ("달서구", "27290"), ("달성군", "27710"),
    ],
}

HEADERS = ["시도", "구군", "법정동", "아파트명", "전용면적(㎡)", "층", "건축년도",
           "계약년", "계약월", "계약일", "거래금액(만원)", "거래유형", "중개사소재지", "수집일시"]


def fetch_data(lawd_cd, deal_ymd, page=1, num_rows=1000):
    params = {
        "serviceKey": API_KEY,
        "LAWD_CD": lawd_cd,
        "DEAL_YMD": deal_ymd,
        "pageNo": page,
        "numOfRows": num_rows,
    }
    try:
        resp = requests.get(API_URL, params=params, timeout=30)
        resp.raise_for_status()
        return resp.text
    except Exception as e:
        print(f"  [오류] {lawd_cd} {deal_ymd} 요청 실패: {e}")
        return None


def parse_items(xml_text):
    try:
        root = ET.fromstring(xml_text)
        items = root.findall(".//item")
        return items
    except Exception as e:
        print(f"  [오류] XML 파싱 실패: {e}")
        return []


def get_text(item, tag, default=""):
    el = item.find(tag)
    return el.text.strip() if el is not None and el.text else default


def fetch_region_data(city_name, districts, deal_ymd):
    all_rows = []
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    for district_name, lawd_cd in districts:
        print(f"  수집 중: {city_name} {district_name} ({deal_ymd})")
        xml = fetch_data(lawd_cd, deal_ymd)
        if not xml:
            continue
        items = parse_items(xml)
        for item in items:
            row = [
                city_name,
                district_name,
                get_text(item, "umdNm"),
                get_text(item, "aptNm"),
                get_text(item, "excluUseAr"),
                get_text(item, "floor"),
                get_text(item, "buildYear"),
                get_text(item, "dealYear"),
                get_text(item, "dealMonth"),
                get_text(item, "dealDay"),
                get_text(item, "dealAmount").replace(",", ""),
                get_text(item, "dealingGbn"),
                get_text(item, "estateAgentSggNm"),
                now_str,
            ]
            all_rows.append(row)
        print(f"    → {len(items)}건 수집")
    return all_rows


def style_header(ws):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF", size=10)
    align = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def load_master():
    existing = {"서울": set(), "대구": set()}
    rows = {"서울": [], "대구": []}
    if not os.path.exists(MASTER_FILE):
        return existing, rows
    wb = openpyxl.load_workbook(MASTER_FILE)
    for city in ["서울", "대구"]:
        if city not in wb.sheetnames:
            continue
        for row in wb[city].iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rows[city].append(list(row))
            key = (str(row[3]), str(row[7]), str(row[8]), str(row[9]), str(row[10]), str(row[5]))
            existing[city].add(key)
    return existing, rows


def merge_and_save(new_seoul, new_daegu):
    existing, rows = load_master()

    added = {"서울": 0, "대구": 0}
    for city, new_rows in [("서울", new_seoul), ("대구", new_daegu)]:
        for row in new_rows:
            key = (str(row[3]), str(row[7]), str(row[8]), str(row[9]), str(row[10]), str(row[5]))
            if key not in existing[city]:
                existing[city].add(key)
                rows[city].append(row)
                added[city] += 1

    for city in ["서울", "대구"]:
        rows[city].sort(key=lambda r: (str(r[7]), str(r[8]).zfill(2), str(r[9]).zfill(2)))

    wb = openpyxl.Workbook()
    for city in ["서울", "대구"]:
        ws = wb.create_sheet(city)
        ws.append(HEADERS)
        for row in rows[city]:
            ws.append(row)
        style_header(ws)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    os.makedirs(os.path.dirname(MASTER_FILE), exist_ok=True)
    wb.save(MASTER_FILE)
    print(f"\n신규 추가: 서울 {added['서울']}건 / 대구 {added['대구']}건")
    print(f"누적 합계: 서울 {len(rows['서울'])}건 / 대구 {len(rows['대구'])}건")
    print(f"저장 완료: {MASTER_FILE}")


def save_to_excel(seoul_rows, daegu_rows, output_path):
    wb = openpyxl.Workbook()

    for sheet_name, rows in [("서울", seoul_rows), ("대구", daegu_rows)]:
        ws = wb.create_sheet(sheet_name)
        ws.append(HEADERS)
        for row in rows:
            ws.append(row)
        style_header(ws)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)
    print(f"일별 스냅샷 저장: {output_path}")


def main():
    now = datetime.now()
    deal_ymd = now.strftime("%Y%m")

    base_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(base_dir, "data")
    os.makedirs(data_dir, exist_ok=True)

    filename = f"아파트실거래_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    output_path = os.path.join(data_dir, filename)

    print(f"=== 아파트 실거래 데이터 수집 시작 ({now.strftime('%Y-%m-%d %H:%M')}) ===")
    print(f"조회 기간: {deal_ymd}\n")

    print("[서울]")
    seoul_rows = fetch_region_data("서울", REGIONS["서울"], deal_ymd)

    print("\n[대구]")
    daegu_rows = fetch_region_data("대구", REGIONS["대구"], deal_ymd)

    print(f"\n총 수집: 서울 {len(seoul_rows)}건 / 대구 {len(daegu_rows)}건")

    save_to_excel(seoul_rows, daegu_rows, output_path)
    merge_and_save(seoul_rows, daegu_rows)


if __name__ == "__main__":
    main()
