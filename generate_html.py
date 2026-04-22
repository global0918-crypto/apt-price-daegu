import openpyxl
import os
import glob
import json

def load_latest_excel():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    master = os.path.join(base_dir, "data", "아파트실거래_마스터.xlsx")
    if os.path.exists(master):
        print(f"마스터 파일 사용: {master}")
        return master
    files = sorted(glob.glob(os.path.join(base_dir, "data", "아파트실거래_*.xlsx")), reverse=True)
    if not files:
        raise FileNotFoundError("data 폴더에 엑셀 파일이 없습니다.")
    return files[0]

def read_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            rows.append(row)
    return rows

def price_label(amount_str):
    try:
        amount = int(str(amount_str).replace(",", ""))
        if amount >= 10000:
            uk = amount // 10000
            rem = amount % 10000
            return f"{uk}억 {rem:,}만원" if rem else f"{uk}억원"
        return f"{amount:,}만원"
    except:
        return f"{amount_str}만원"

def price_tier(amount_str):
    try:
        amount = int(str(amount_str).replace(",", ""))
        if amount >= 100000: return "tier-high"
        if amount >= 50000:  return "tier-mid"
        return "tier-low"
    except:
        return "tier-low"

def make_cards(rows, hgn=None):
    hgn = hgn or {}
    cards = []
    for r in rows:
        sido, gugun, dong, apt_name, area, floor, build_year, deal_year, deal_month, deal_day, amount, deal_type, agent_loc, collected = r
        try:
            area_f = float(area)
            pyeong = round(area_f * 1.32 / 3.305785)
            area_str = f"전용 {round(area_f)}㎡({pyeong}평)"
            area_key = f"{round(area_f)}㎡"
        except:
            area_f = 0
            area_str = f"전용 {area}㎡"
            area_key = f"{area}㎡"

        tier = price_tier(amount)
        price_str = price_label(amount)
        date_str = f"{deal_year}.{str(deal_month).zfill(2)}.{str(deal_day).zfill(2)}"

        try:
            amount_raw = int(str(amount).replace(",", ""))
        except:
            amount_raw = 0

        hgn_info = hgn.get(apt_name, {})
        review_count = hgn_info.get("count", 0)
        hgn_hash = hgn_info.get("hash", None)

        cards.append({
            "sido": sido, "gugun": gugun, "dong": dong,
            "apt_name": apt_name, "area": area_str, "area_key": area_key,
            "area_f": area_f, "floor": floor, "build_year": build_year,
            "date": date_str, "price": price_str,
            "deal_type": deal_type or "매매", "tier": tier,
            "amount_raw": amount_raw,
            "review_count": review_count,
            "hgn_hash": hgn_hash,
            "agent_loc": agent_loc or "",
            "collected": str(collected) if collected else "",
        })
    return cards

def load_hogangnono():
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "hogangnono_reviews.json")
    if not os.path.exists(path):
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def load_latest_snapshot():
    """마스터 파일 외 가장 최근 일별 스냅샷 반환 (카드용)"""
    base_dir = os.path.dirname(os.path.abspath(__file__))
    files = sorted(glob.glob(os.path.join(base_dir, "data", "아파트실거래_2*.xlsx")), reverse=True)
    if not files:
        return None
    return files[0]

def build_spark_map(all_rows):
    """7년 전체 이력 → {아파트명: [{d,v,ak,f,dt,g}]} (Python 미리 계산)
    g = gugun (구군) — 동명 아파트 구분용
    """
    from collections import defaultdict
    m = defaultdict(list)
    for r in all_rows:
        _, gugun, _, apt_name, area, floor, _, deal_year, deal_month, deal_day, amount, deal_type, *_ = r
        try:
            v = int(str(amount).replace(",", ""))
            d = f"{deal_year}.{str(deal_month).zfill(2)}.{str(deal_day).zfill(2)}"
            try:
                ak = f"{round(float(area))}㎡"
            except:
                ak = f"{area}㎡"
            m[apt_name].append({"d": d, "v": v, "ak": ak, "f": str(floor) if floor else "", "dt": str(deal_type) if deal_type else "매매", "g": str(gugun) if gugun else ""})
        except:
            pass
    return {k: sorted(v, key=lambda x: x["d"]) for k, v in m.items()}

def render_cards_js(cards):
    return json.dumps(cards, ensure_ascii=False)

def generate_html(excel_path, output_path):
    base_dir = os.path.dirname(os.path.abspath(__file__))
    hgn = load_hogangnono()

    # ── SPARK_MAP: 마스터(7년) 전체로 계산 (실패 시 스냅샷으로 폴백) ──
    master_rows = []
    try:
        wb_master = openpyxl.load_workbook(excel_path)
        master_rows = read_sheet(wb_master, "대구")  # 대구만
        print(f"마스터 로드: 대구 {len(master_rows)}건")
    except Exception as e:
        print(f"마스터 파일 로드 실패 ({e}), 스냅샷으로 폴백")

    # ── CARDS: 최신 스냅샷(당월)만 카드에 표시 ──
    snap_path = load_latest_snapshot()
    if snap_path:
        try:
            wb_snap = openpyxl.load_workbook(snap_path)
            snap_rows = read_sheet(wb_snap, "대구")  # 대구만
            print(f"카드용 스냅샷: {snap_path} ({len(snap_rows)}건)")
            if not master_rows:
                master_rows = snap_rows
        except Exception as e:
            print(f"스냅샷 로드 실패: {e}")
            snap_rows = master_rows
    else:
        snap_rows = master_rows

    spark_map = build_spark_map(master_rows)
    spark_json = json.dumps(spark_map, ensure_ascii=False, separators=(',', ':'))
    print(f"SPARK_MAP: {len(spark_map)}개 아파트, {sum(len(v) for v in spark_map.values())}건")
    cards = make_cards(snap_rows, hgn)

    from datetime import datetime
    date_label = datetime.now().strftime("%Y%m%d_%H%M")
    cards_json = render_cards_js(cards)

    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>대구 아파트 실거래 현황</title>

<!-- Open Graph -->
<meta property="og:type"               content="website">
<meta property="og:title"              content="대구 아파트 실거래 현황">
<meta property="og:description"        content="국토교통부 실거래가 기반 대구 아파트 실거래 카드뷰 · 7년 추이 · 유사 아파트 비교">
<meta property="og:image"              content="apt_price_daegu_thumbnail.png">
<meta property="og:image:width"        content="1200">
<meta property="og:image:height"       content="630">
<meta name="twitter:card"              content="summary_large_image">
<meta name="twitter:image"             content="apt_price_daegu_thumbnail.png">

<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, 'Noto Sans KR', sans-serif; background: #f4f4f4; color: #111; }}

  header {{
    background: #03c75a; color: #fff;
    padding: 18px 24px;
    box-shadow: 0 1px 4px rgba(0,0,0,.18);
  }}
  header h1 {{ font-size: 20px; font-weight: 700; letter-spacing: -0.5px; }}
  header p {{ font-size: 12px; color: rgba(255,255,255,0.75); margin-top: 4px; }}

  .toolbar {{
    background: #fff; padding: 12px 20px;
    display: flex; flex-direction: column; gap: 0;
    border-bottom: 1px solid #e4e4e4;
    position: sticky; top: 0; z-index: 100;
    box-shadow: 0 1px 4px rgba(0,0,0,.06);
  }}
  .toolbar-row {{ display: flex; flex-wrap: wrap; gap: 8px; align-items: center; }}
  .toolbar-row.gugun-row {{
    padding-bottom: 10px; margin-bottom: 10px;
    border-bottom: 1px solid #f0f0f0;
  }}
  .toolbar input, .toolbar select {{
    border: 1px solid #ddd; border-radius: 6px;
    padding: 7px 12px; font-size: 13px; outline: none;
    background: #fafafa; transition: border .15s; color: #111;
  }}
  .toolbar input:focus, .toolbar select:focus {{ border-color: #03c75a; background: #fff; }}
  .search-input {{ width: 180px; }}
  .count-badge {{
    margin-left: auto; background: #f0f0f0; border-radius: 20px;
    padding: 5px 14px; font-size: 13px; color: #555; font-weight: 500;
  }}
  .count-badge span {{ color: #111; font-weight: 700; }}

  /* 구군 칩 */
  .gugun-chips {{ display: flex; flex-wrap: wrap; gap: 5px; }}
  .gugun-chip {{
    padding: 4px 12px; border-radius: 16px; border: 1px solid #ddd;
    background: #fafafa; font-size: 12px; font-weight: 500; cursor: pointer;
    transition: all .15s; white-space: nowrap; color: #555;
  }}
  .gugun-chip.active {{ background: #03c75a; color: #fff; border-color: #03c75a; }}
  .gugun-chip:hover:not(.active) {{ border-color: #03c75a; color: #03c75a; }}

  /* 날짜 기간 */
  .date-range {{ display: flex; align-items: center; gap: 6px; flex-wrap: wrap; }}
  .date-range input[type=date] {{
    padding: 6px 10px; border: 1px solid #ddd; border-radius: 6px;
    font-size: 12px; outline: none; background: #fafafa; color: #333;
  }}
  .date-range input[type=date]:focus {{ border-color: #03c75a; }}
  .date-sep {{ color: #bbb; font-size: 13px; }}
  .quick-dates {{ display: flex; gap: 4px; }}
  .quick-btn {{
    padding: 5px 10px; border: 1px solid #ddd; border-radius: 6px;
    font-size: 12px; font-weight: 500; background: #fafafa; cursor: pointer;
    color: #555; transition: all .15s;
  }}
  .quick-btn:hover {{ background: #03c75a; color: #fff; border-color: #03c75a; }}
  .latest-date {{ font-size: 11px; color: #bbb; padding-left: 4px; }}

  .grid {{
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
    gap: 14px; padding: 20px 24px;
    max-width: 1600px; margin: 0 auto;
  }}

  .card {{
    background: #fff; border-radius: 12px;
    border: 1px solid #e8e8e8;
    box-shadow: 0 1px 4px rgba(0,0,0,.05);
    overflow: hidden; transition: box-shadow .18s;
    cursor: pointer;
  }}
  .card:hover {{ box-shadow: 0 4px 16px rgba(0,0,0,.11); }}
  .card:hover .apt-name {{ color: #333; }}

  .card-header {{ padding: 12px 16px 10px; border-bottom: 1px solid #f0f0f0; }}
  .card-header-top {{
    display: flex; justify-content: space-between; align-items: flex-start; gap: 6px;
    margin-bottom: 3px;
  }}
  .apt-name {{
    font-size: 14px; font-weight: 700; color: #111;
    white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
    transition: color .15s; flex: 1; min-width: 0;
  }}
  .review-badge {{
    flex-shrink: 0; display: flex; align-items: center; gap: 3px;
    background: #f5f5f5; border-radius: 8px; padding: 2px 8px;
    font-size: 11px; color: #888; font-weight: 500; white-space: nowrap;
    text-decoration: none; border: 1px solid #eee;
  }}
  .review-badge:hover {{ background: #eee; color: #444; }}
  .review-badge.zero {{ color: #ccc; }}
  .location {{ font-size: 11px; color: #999; }}

  .card-body {{ padding: 12px 16px 14px; }}

  .price-spark-row {{
    display: flex; align-items: center; justify-content: space-between;
    gap: 10px; margin-bottom: 2px;
  }}
  .price {{ font-size: 18px; font-weight: 800; letter-spacing: -0.5px; flex-shrink: 0; color: #111; }}
  .tier-high .price {{ color: #c0392b; }}
  .tier-mid  .price {{ color: #333; }}
  .tier-low  .price {{ color: #555; }}
  .sparkline-wrap {{
    flex-shrink: 0; display: flex; align-items: center;
  }}
  .sparkline-wrap svg {{ display: block; overflow: visible; }}

  .alltime-high {{
    font-size: 11px; color: #bbb; margin-bottom: 10px;
    display: flex; align-items: center; gap: 4px;
  }}
  .alltime-high .ath-val {{ color: #555; font-weight: 600; }}
  .alltime-high .ath-vs.minus {{ color: #03c75a; }}
  .alltime-high .ath-vs.plus {{ color: #c0392b; }}

  .deal-badge {{
    font-size: 11px; padding: 2px 8px; border-radius: 8px;
    background: #f5f5f5; color: #777; font-weight: 400; border: 1px solid #eee;
  }}
  .info-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 8px; }}
  .info-item {{ display: flex; flex-direction: column; gap: 2px; }}
  .info-label {{ font-size: 11px; color: #bbb; }}
  .info-value {{ font-size: 13px; color: #333; font-weight: 500; }}

  .card-footer {{
    padding: 8px 16px; background: #fafafa;
    border-top: 1px solid #f0f0f0;
    display: flex; justify-content: space-between; align-items: center;
  }}
  .chart-hint {{ font-size: 11px; color: #999; }}
  .dl-btn {{
    display: inline-flex; align-items: center; gap: 5px;
    padding: 6px 14px; border: 1px solid #03c75a; border-radius: 6px;
    background: #fff; color: #03c75a; font-size: 13px; font-weight: 600;
    cursor: pointer; transition: all .15s; white-space: nowrap;
  }}
  .dl-btn:hover {{ background: #03c75a; color: #fff; }}

  .empty {{ text-align: center; padding: 80px 24px; color: #bbb; font-size: 15px; grid-column: 1/-1; }}

  /* ── 모달 ── */
  .modal-overlay {{
    display: none; position: fixed; inset: 0;
    background: rgba(0,0,0,.45); z-index: 1000;
    align-items: center; justify-content: center;
  }}
  .modal-overlay.open {{ display: flex; }}
  .modal {{
    background: #fff; border-radius: 16px;
    width: min(820px, 95vw); max-height: 92vh;
    display: flex; flex-direction: column;
    box-shadow: 0 16px 48px rgba(0,0,0,.2);
    animation: slideUp .2s ease;
  }}
  @keyframes slideUp {{
    from {{ transform: translateY(24px); opacity: 0; }}
    to   {{ transform: translateY(0);    opacity: 1; }}
  }}
  .modal-header {{
    padding: 20px 24px 0; flex-shrink: 0;
    display: flex; align-items: flex-start; justify-content: space-between;
  }}
  .modal-title {{ font-size: 17px; font-weight: 700; color: #111; }}
  .modal-sub   {{ font-size: 12px; color: #999; margin-top: 3px; }}
  .modal-close {{
    background: none; border: none; font-size: 20px; cursor: pointer;
    color: #bbb; padding: 0 4px; line-height: 1; flex-shrink: 0;
  }}
  .modal-close:hover {{ color: #333; }}

  /* ── 메인 탭 ── */
  .main-tabs {{
    display: flex; gap: 0; padding: 14px 24px 0; flex-shrink: 0;
    border-bottom: 1px solid #eee;
  }}
  .main-tab {{
    padding: 8px 20px; font-size: 13px; font-weight: 600;
    border: none; background: none; cursor: pointer; color: #bbb;
    border-bottom: 2px solid transparent; margin-bottom: -1px;
    transition: color .15s, border-color .15s;
  }}
  .main-tab.active {{ color: #111; border-bottom-color: #111; }}
  .main-tab:hover:not(.active) {{ color: #555; }}

  .tab-panel {{ display: none; }}
  .tab-panel.active {{ display: block; }}
  .modal-body {{ padding: 16px 24px; overflow-y: auto; }}
  .chart-wrap {{ position: relative; height: 300px; margin-bottom: 18px; }}

  /* ── 면적 타입 탭 ── */
  .type-tabs {{ display: flex; flex-wrap: wrap; gap: 5px; margin-bottom: 12px; }}
  .type-tab {{
    padding: 4px 12px; border-radius: 20px; font-size: 12px; font-weight: 500;
    border: 1px solid #ddd; cursor: pointer; transition: all .15s;
    background: #fafafa; color: #666;
  }}
  .type-tab.active {{ background: #03c75a; color: #fff; border-color: #03c75a; }}
  .type-tab:hover:not(.active) {{ border-color: #555; color: #111; }}

  .chart-meta {{
    display: grid; grid-template-columns: repeat(3, 1fr);
    gap: 10px; margin-bottom: 16px;
  }}
  .meta-box {{
    background: #f8f8f8; border-radius: 8px;
    padding: 10px 14px; text-align: center;
  }}
  .meta-label {{ font-size: 11px; color: #bbb; margin-bottom: 3px; }}
  .meta-value {{ font-size: 14px; font-weight: 700; color: #111; }}
  .meta-value.up   {{ color: #c0392b; }}
  .meta-value.down {{ color: #03c75a; }}

  .trade-table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  .trade-table th {{
    background: #f8fafb; padding: 8px 12px;
    text-align: left; color: #888; font-weight: 500;
    border-bottom: 1px solid #eee; position: sticky; top: 0;
  }}
  .trade-table td {{ padding: 8px 12px; border-bottom: 1px solid #f0f0f0; color: #333; }}
  .trade-table tr:last-child td {{ border-bottom: none; }}
  .trade-table tr:hover td {{ background: #fafafa; }}
  .price-cell {{ font-weight: 700; color: #111; }}
  .price-cell.high {{ color: #c0392b; }}
  .price-cell.low  {{ color: #03c75a; }}
  .dot {{ display: inline-block; width: 8px; height: 8px; border-radius: 50%; margin-right: 5px; vertical-align: middle; }}

  /* ── 비교 탭 ── */
  .cmp-legend {{ display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 14px; }}
  .cmp-legend-item {{
    display: flex; align-items: center; gap: 5px;
    font-size: 12px; color: #555; font-weight: 500;
  }}
  .cmp-dot {{ width: 8px; height: 8px; border-radius: 50%; flex-shrink: 0; }}
  .cmp-self {{ font-weight: 700; color: #111; }}
  .cmp-table {{ width: 100%; border-collapse: collapse; font-size: 12px; margin-top: 16px; }}
  .cmp-table th {{
    background: #f8f8f8; padding: 8px 10px; text-align: left;
    color: #999; font-weight: 500; border-bottom: 1px solid #eee;
  }}
  .cmp-table td {{ padding: 8px 10px; border-bottom: 1px solid #f0f0f0; }}
  .cmp-table tr:last-child td {{ border-bottom: none; }}
  .cmp-self-row td {{ font-weight: 700; background: #f8f8f8; }}

  /* ── 모바일 최적화 ── */
  @media (max-width: 640px) {{
    header {{ padding: 14px 16px; }}
    header h1 {{ font-size: 18px; }}
    .toolbar {{ padding: 10px 14px; }}
    .search-input {{ width: 140px; }}
    .grid {{ grid-template-columns: 1fr; padding: 12px 14px; gap: 12px; }}
    .card-body {{ padding: 10px 14px 12px; }}
    .price {{ font-size: 16px; }}
    .modal {{
      width: 100%; max-height: 92vh; border-radius: 18px 18px 0 0;
      position: fixed; bottom: 0; margin: 0;
    }}
    .modal-overlay.open {{ align-items: flex-end; }}
    .chart-wrap {{ height: 220px; }}
    .main-tab {{ padding: 8px 14px; font-size: 13px; }}
    .chart-meta {{ grid-template-columns: 1fr 1fr; }}
    .trade-table th, .trade-table td {{ padding: 7px 8px; font-size: 12px; }}
  }}
  @media (max-width: 400px) {{
    .grid {{ grid-template-columns: 1fr; padding: 10px; }}
    .info-grid {{ grid-template-columns: 1fr 1fr; gap: 6px; }}
    .quick-dates {{ gap: 3px; }}
    .quick-btn {{ padding: 4px 8px; font-size: 11px; }}
  }}

  /* ── 신고가 카드 ── */
  .new-high {{ border: 2px solid #e74c3c !important; box-shadow: 0 0 0 4px rgba(231,76,60,.1), 0 1px 4px rgba(0,0,0,.05); }}
  .new-high-badge {{ display: inline-block; background: #e74c3c; color: #fff; font-size: 10px; font-weight: 700; padding: 2px 6px; border-radius: 5px; margin-left: 5px; vertical-align: middle; }}

  /* ── 랜딩 오버레이 ── */
  .landing-overlay {{ position: fixed; inset: 0; background: #f4f4f4; z-index: 2000; display: flex; flex-direction: column; overflow-y: auto; }}
  .landing-inner {{ max-width: 680px; width: 100%; margin: 0 auto; padding: 40px 24px 60px; }}
  .landing-hd {{ text-align: center; margin-bottom: 36px; }}
  .landing-hd h1 {{ font-size: 22px; font-weight: 800; color: #111; margin-bottom: 6px; }}
  .landing-hd p {{ font-size: 13px; color: #888; }}
  .landing-sec {{ margin-bottom: 28px; }}
  .landing-sec-title {{ font-size: 13px; font-weight: 700; color: #888; margin-bottom: 12px; text-transform: uppercase; letter-spacing: .3px; }}
  .ld-date-chips {{ display: flex; flex-wrap: wrap; gap: 6px; }}
  .ld-date-chip {{
    padding: 9px 13px; border-radius: 10px; border: 1.5px solid #ddd;
    background: #fff; font-size: 13px; font-weight: 500; cursor: pointer;
    transition: all .15s; color: #444; text-align: center; line-height: 1.3; min-width: 58px;
  }}
  .ld-date-chip .chip-sub {{ font-size: 10px; color: #bbb; display: block; margin-top: 1px; }}
  .ld-date-chip.today {{ border-color: #333; font-weight: 700; }}
  .ld-date-chip.today .chip-sub {{ color: #666; }}
  .ld-date-chip.active {{ background: #03c75a; color: #fff; border-color: #03c75a; }}
  .ld-date-chip.active .chip-sub {{ color: rgba(255,255,255,.7); }}
  .ld-date-chip:hover:not(.active):not([disabled]) {{ border-color: #03c75a; color: #03c75a; }}
  .ld-date-chip[disabled] {{ background: #f5f5f5; color: #ccc; border-color: #ebebeb; cursor: not-allowed; pointer-events: none; }}
  .ld-date-chip[disabled] .chip-sub {{ color: #ccc; }}
  .ld-gugun-chips {{ display: flex; flex-wrap: wrap; gap: 7px; }}
  .ld-gugun-chip {{
    padding: 9px 20px; border-radius: 24px; border: 1.5px solid #ddd;
    background: #fff; font-size: 13px; font-weight: 500; cursor: pointer;
    transition: all .15s; color: #555;
  }}
  .ld-gugun-chip.active {{ background: #03c75a; color: #fff; border-color: #03c75a; }}
  .ld-gugun-chip:hover:not(.active) {{ border-color: #03c75a; color: #03c75a; }}
  .ld-info {{ text-align: center; font-size: 12px; color: #aaa; margin-top: 12px; }}
  .ld-info b {{ color: #333; }}
  .ld-submit {{
    width: 100%; padding: 16px; background: #03c75a; color: #fff;
    border: none; border-radius: 12px; font-size: 16px; font-weight: 700;
    cursor: pointer; margin-top: 8px; transition: background .15s;
    box-shadow: 0 4px 18px rgba(3,199,90,.3);
  }}
  .ld-submit:hover {{ background: #02a84b; }}
  @media (max-width: 640px) {{
    .landing-inner {{ padding: 28px 16px 48px; }}
    .landing-hd h1 {{ font-size: 19px; }}
    .ld-date-chip {{ padding: 8px 10px; font-size: 12px; min-width: 50px; }}
    .ld-gugun-chip {{ padding: 8px 14px; font-size: 12px; }}
  }}
</style>
</head>
<body>

<!-- 랜딩 오버레이 -->
<div id="landingOverlay" class="landing-overlay">
  <div class="landing-inner">
    <div class="landing-hd">
      <h1>🏢 대구 아파트 실거래 현황</h1>
      <p>국토교통부 실거래가 공개시스템</p>
    </div>
    <div class="landing-sec">
      <div class="landing-sec-title">📅 조회 날짜</div>
      <div class="ld-date-chips" id="ldDateChips"></div>
      <div class="ld-info" id="ldInfo"></div>
    </div>
    <div class="landing-sec">
      <div class="landing-sec-title">📍 구군 선택</div>
      <div class="ld-gugun-chips" id="ldGugunChips"></div>
    </div>
    <button class="ld-submit" onclick="applyLanding()">조회하기</button>
  </div>
</div>

<header>
  <h1>🏢 아파트 실거래 현황</h1>
  <p>국토교통부 실거래가 공개시스템 · 기준일: {date_label[:8]}</p>
</header>

<div class="toolbar">
  <!-- 행1: 구군 칩 -->
  <div class="toolbar-row gugun-row">
    <div class="gugun-chips" id="gugunChips"></div>
  </div>
  <!-- 행2: 검색 + 날짜 + 정렬 + 카운트 -->
  <div class="toolbar-row">
    <input type="text" id="searchApt" class="search-input" placeholder="아파트명 검색..." oninput="applyFilter()">
    <div class="date-range">
      <input type="date" id="dateFrom" onchange="applyFilter()">
      <span class="date-sep">~</span>
      <input type="date" id="dateTo" onchange="applyFilter()">
      <div class="quick-dates">
        <button class="quick-btn" onclick="setQuickDate(1)">1일</button>
        <button class="quick-btn" onclick="setQuickDate(7)">일주일</button>
        <button class="quick-btn" onclick="setQuickDate(30)">한달</button>
      </div>
      <button class="quick-btn" onclick="applyFilter()" style="background:#03c75a;color:#fff;border-color:#03c75a;">조회</button>
      <button class="quick-btn" onclick="resetFilter()">초기화</button>
      <span class="latest-date" id="latestDateLabel"></span>
    </div>
    <select id="selSort" onchange="applyFilter()">
      <option value="date_desc">최신순</option>
      <option value="price_desc">가격 높은순</option>
      <option value="price_asc">가격 낮은순</option>
      <option value="area_desc">면적 넓은순</option>
    </select>
    <div class="count-badge">총 <span id="countNum">0</span>건</div>
    <button class="dl-btn" onclick="downloadExcel()">⬇ 엑셀 다운로드</button>
  </div>
</div>

<div class="grid" id="grid"></div>
<div style="text-align:center;padding:20px 0 40px;">
  <button id="loadMoreBtn" onclick="loadMore()" style="display:none;background:#03c75a;color:white;border:none;border-radius:8px;padding:12px 32px;font-size:15px;font-weight:600;cursor:pointer;box-shadow:0 2px 8px rgba(3,199,90,.3);">더보기</button>
</div>

<!-- 모달 -->
<div class="modal-overlay" id="modalOverlay" onclick="closeModal(event)">
  <div class="modal" id="modal">
    <div class="modal-header">
      <div>
        <div class="modal-title" id="modalTitle"></div>
        <div class="modal-sub"  id="modalSub"></div>
      </div>
      <button class="modal-close" onclick="closeModalDirect()">✕</button>
    </div>

    <!-- 메인 탭 -->
    <div class="main-tabs">
      <button class="main-tab active" onclick="switchMainTab('trend', this)">📈 실거래가 추이</button>
      <button class="main-tab"       onclick="switchMainTab('compare', this)">🏘 유사 아파트 비교</button>
    </div>

    <div class="modal-body">
      <!-- 탭1: 실거래가 추이 -->
      <div class="tab-panel active" id="panel-trend">
        <div class="type-tabs" id="typeTabs"></div>
        <div class="chart-meta" id="chartMeta"></div>
        <div class="chart-wrap"><canvas id="priceChart"></canvas></div>
        <table class="trade-table">
          <thead><tr><th>계약일</th><th>타입</th><th>층</th><th>거래금액</th><th>거래유형</th></tr></thead>
          <tbody id="tradeBody"></tbody>
        </table>
      </div>

      <!-- 탭2: 유사 아파트 비교 -->
      <div class="tab-panel" id="panel-compare">
        <div class="cmp-legend" id="cmpLegend"></div>
        <div class="chart-wrap"><canvas id="cmpChart"></canvas></div>
        <table class="cmp-table">
          <thead><tr><th>아파트</th><th>평균가</th><th>최고가</th><th>최저가</th><th>거래수</th></tr></thead>
          <tbody id="cmpBody"></tbody>
        </table>
      </div>
    </div>
  </div>
</div>

<script>
const DATA = {cards_json};
const SPARK_MAP = {spark_json};
let chartInst = null;

/* ── 초기화 ── */
let _activeGugun = '';

function buildSelects() {{
  // 구군 칩
  const guguns = [...new Set(DATA.map(d => d.gugun))].sort();
  const chips = document.getElementById('gugunChips');
  chips.innerHTML = `<button class="gugun-chip active" data-g="" onclick="selectGugun(this,'')">전체</button>`;
  guguns.forEach(g => {{
    chips.innerHTML += `<button class="gugun-chip" data-g="${{g}}" onclick="selectGugun(this,'${{g}}')">${{g}}</button>`;
  }});

  // 날짜 범위 초기화: 가장 최근 계약일 기준 한달
  const dates = DATA.map(d => d.date).filter(Boolean).sort();
  const latestDate = dates[dates.length - 1] || '';
  if (latestDate) {{
    /* 기본: 가장 최근 계약일 기준 1일 */
    const isoLatest = latestDate.replace(/\./g, '-');
    document.getElementById('dateTo').value = isoLatest;
    document.getElementById('dateFrom').value = isoLatest;
    document.getElementById('latestDateLabel').textContent = `최근 ${{latestDate}}`;
  }}
}}

function selectGugun(el, g) {{
  _activeGugun = g;
  document.querySelectorAll('.gugun-chip').forEach(c => c.classList.remove('active'));
  el.classList.add('active');
  applyFilter();
}}

function setQuickDate(days) {{
  const dates = DATA.map(d => d.date).filter(Boolean).sort();
  const latest = dates[dates.length - 1] || '';
  if (!latest) return;
  document.getElementById('dateTo').value = latest.replace(/\./g, '-');
  const dt = new Date(latest.replace(/\./g, '-'));
  dt.setDate(dt.getDate() - (days - 1));
  document.getElementById('dateFrom').value = dt.toISOString().slice(0,10);
  applyFilter();
}}

function resetFilter() {{
  _activeGugun = '';
  document.querySelectorAll('.gugun-chip').forEach(c => c.classList.remove('active'));
  document.querySelector('.gugun-chip[data-g=""]')?.classList.add('active');
  document.getElementById('searchApt').value = '';
  document.getElementById('dateFrom').value = '';
  document.getElementById('dateTo').value = '';
  document.getElementById('selSort').value = 'date_desc';
  applyFilter();
}}

/* ── 필터 & 정렬 ── */
function applyFilter() {{
  const keyword  = document.getElementById('searchApt').value.trim().toLowerCase();
  const sort     = document.getElementById('selSort').value;
  /* date input은 "YYYY-MM-DD", DATA.date는 "YYYY.MM.DD" → 점으로 통일 */
  const dateFrom = document.getElementById('dateFrom').value.replace(/-/g,'.');
  const dateTo   = document.getElementById('dateTo').value.replace(/-/g,'.');

  let filtered = DATA.filter(d => {{
    if (_activeGugun && d.gugun !== _activeGugun) return false;
    if (keyword && !d.apt_name.toLowerCase().includes(keyword)) return false;
    if (dateFrom && d.date < dateFrom) return false;
    if (dateTo   && d.date > dateTo)   return false;
    return true;
  }});
  filtered.sort((a, b) => {{
    if (sort === 'price_desc') return b.amount_raw - a.amount_raw;
    if (sort === 'price_asc')  return a.amount_raw - b.amount_raw;
    if (sort === 'date_desc')  return b.date.localeCompare(a.date);
    if (sort === 'area_desc')  return b.area_f - a.area_f;
    return 0;
  }});
  renderCards(filtered);
}}

/* ── 카드 렌더 (더보기 페이지네이션) ── */
let _currentList = [];
let _visibleCount = 60;
const PAGE_SIZE = 60;

function renderCards(list) {{
  _currentList = list;
  _visibleCount = PAGE_SIZE;
  document.getElementById('countNum').textContent = list.length.toLocaleString();
  const grid = document.getElementById('grid');
  if (!list.length) {{
    grid.innerHTML = '<div class="empty">검색 결과가 없습니다.</div>';
    document.getElementById('loadMoreBtn').style.display = 'none';
    return;
  }}
  renderVisible();
}}

function renderVisible() {{
  const list = _currentList;
  const slice = list.slice(0, _visibleCount);
  const grid = document.getElementById('grid');
  grid.innerHTML = slice.map((d, i) => {{
    const allTrades  = (SPARK_MAP[d.apt_name] || []).filter(t => t.ak === d.area_key && t.g === d.gugun);
    const allPrices  = allTrades.map(t => t.v).filter(v => v > 0);
    const athVal     = allPrices.length ? Math.max(...allPrices) : 0;
    const athDiff    = athVal ? d.amount_raw - athVal : 0;
    const prevPrices = allTrades.filter(t => t.d < d.date).map(t => t.v).filter(v => v > 0);
    const prevAth    = prevPrices.length ? Math.max(...prevPrices) : 0;
    const isNewHigh  = prevAth > 0 && d.amount_raw > prevAth;
    const athDiffStr = athDiff === 0 ? '전고점' :
                       (athDiff > 0 ? `+${{fmtPriceMini(athDiff)}}` : `${{fmtPriceMini(athDiff)}}`);
    const athClass   = athDiff > 0 ? 'plus' : athDiff < 0 ? 'minus' : '';
    const athLine    = athVal ? `<div class="alltime-high">
      전고점 <span class="ath-val">${{fmtPriceMini(athVal)}}</span>
      <span class="ath-vs ${{athClass}}">${{athDiff===0?'(현재=전고점)':'('+athDiffStr+')'}}</span>
    </div>` : '';

    const rc = d.review_count || 0;
    const hgnHref = d.hgn_hash ? `https://hogangnono.com/apt/${{d.hgn_hash}}/0` : '#';
    const reviewBadge = `<a class="review-badge ${{rc===0?'zero':''}}" href="${{hgnHref}}" target="_blank" onclick="event.stopPropagation()">
      💬 ${{rc > 0 ? rc.toLocaleString() : '-'}}
    </a>`;

    return `
    <div class="card ${{d.tier}}${{isNewHigh ? ' new-high' : ''}}" onclick="openModal('${{encodeURIComponent(d.apt_name)}}','${{encodeURIComponent(d.gugun)}}','${{encodeURIComponent(d.area_key)}}','${{encodeURIComponent(d.date)}}','${{d.amount_raw}}')">
      <div class="card-header">
        <div class="card-header-top">
          <div class="apt-name">${{d.apt_name}}${{isNewHigh ? '<span class="new-high-badge">신고가</span>' : ''}}</div>
          ${{reviewBadge}}
        </div>
        <div class="location">${{d.sido}} ${{d.gugun}} ${{d.dong}}</div>
      </div>
      <div class="card-body">
        <div class="price-spark-row">
          <div class="price">${{d.price}}</div>
          <div class="sparkline-wrap">${{makeSpark(d.apt_name, d.gugun)}}</div>
        </div>
        ${{athLine}}
        <div class="info-grid">
          <div class="info-item">
            <span class="info-label">전용면적</span>
            <span class="info-value">${{d.area}}</span>
          </div>
          <div class="info-item">
            <span class="info-label">층</span>
            <span class="info-value">${{d.floor}}층</span>
          </div>
          <div class="info-item">
            <span class="info-label">건축년도</span>
            <span class="info-value">${{d.build_year}}년</span>
          </div>
          <div class="info-item">
            <span class="info-label">계약일</span>
            <span class="info-value">${{d.date}}</span>
          </div>
        </div>
      </div>
      <div class="card-footer">
        <span class="deal-badge">${{d.deal_type || '매매'}}</span>
        <span class="chart-hint">📈 클릭하여 추이 보기</span>
      </div>
    </div>`;
  }}).join('');

  const btn = document.getElementById('loadMoreBtn');
  if (_visibleCount >= list.length) {{
    btn.style.display = 'none';
  }} else {{
    btn.style.display = 'block';
    btn.textContent = `더보기 (${{Math.min(PAGE_SIZE, list.length - _visibleCount)}}건 / 남은 ${{list.length - _visibleCount}}건)`;
  }}
}}

function loadMore() {{
  _visibleCount += PAGE_SIZE;
  renderVisible();
}}

function makeSpark(aptName, gugun='', W=100, H=36) {{
  const raw = SPARK_MAP[aptName] || [];
  const pts = gugun ? raw.filter(p => p.g === gugun) : raw;
  if (!pts || pts.length < 2) {{
    return `<svg width="${{W}}" height="${{H}}"><line x1="0" y1="${{H/2}}" x2="${{W}}" y2="${{H/2}}" stroke="#ddd" stroke-width="1.5"/></svg>`;
  }}
  const vals = pts.map(p => p.v);
  const min = Math.min(...vals), max = Math.max(...vals);
  const range = max - min || 1;
  const pad = 3;
  const coords = pts.map((p, i) => {{
    const x = pad + (i / (pts.length - 1)) * (W - pad*2);
    const y = pad + (1 - (p.v - min) / range) * (H - pad*2);
    return `${{x.toFixed(1)}},${{y.toFixed(1)}}`;
  }});
  const last = vals[vals.length-1];
  const color = '#03c75a';
  const fillPts = [
    `${{pad}},${{H}}`, ...coords, `${{(W-pad).toFixed(1)}},${{H}}`
  ].join(' ');
  const lastX = (pad + (W-pad*2)).toFixed(1);
  const lastY = (pad + (1 - (last-min)/range) * (H-pad*2)).toFixed(1);
  return `<svg width="${{W}}" height="${{H}}" viewBox="0 0 ${{W}} ${{H}}">
    <polygon points="${{fillPts}}" fill="${{color}}22"/>
    <polyline points="${{coords.join(' ')}}" fill="none" stroke="${{color}}" stroke-width="1.8" stroke-linejoin="round" stroke-linecap="round"/>
    <circle cx="${{lastX}}" cy="${{lastY}}" r="2.5" fill="${{color}}" stroke="white" stroke-width="1"/>
  </svg>`;
}}

/* ── 팔레트 ── */
const PALETTE = [
  '#03c75a','#2563eb','#f5a623','#d0021b','#8b5cf6',
  '#06b6d4','#f97316','#ec4899','#84cc16','#6366f1',
];

/* ── 면적 포맷 "전용 84㎡(25평)" ── */
function fmtArea(ak) {{
  if (!ak) return '';
  const m2 = parseFloat(ak);
  if (!m2) return ak;
  const pyeong = Math.round(m2 * 1.32 / 3.305785);
  return `전용 ${{Math.round(m2)}}㎡(${{pyeong}}평)`;
}}

let currentAptName    = '';
let currentGugun      = '';
let currentActiveType = '전체';
let currentTradeDate  = '';
let currentTradePrice = 0;

/* ── 모달 열기 ── */
function openModal(encodedName, encodedGugun, encodedAreaKey, encodedDate, tradePrice) {{
  currentAptName    = decodeURIComponent(encodedName);
  currentGugun      = encodedGugun ? decodeURIComponent(encodedGugun) : '';
  currentTradeDate  = encodedDate  ? decodeURIComponent(encodedDate)  : '';
  currentTradePrice = tradePrice   ? Number(tradePrice)               : 0;
  const defaultType = encodedAreaKey ? decodeURIComponent(encodedAreaKey) : '전체';
  currentActiveType = defaultType;

  /* SPARK_MAP(7년치)에서 해당 구군 데이터만 추출 */
  const allRaw = SPARK_MAP[currentAptName] || [];
  const all = currentGugun ? allRaw.filter(p => p.g === currentGugun) : allRaw;
  if (!all.length) return;

  /* 카드에서 위치·건축연도 정보 가져오기 (같은 구군 우선) */
  const card = DATA.find(d => d.apt_name === currentAptName && d.gugun === currentGugun)
            || DATA.find(d => d.apt_name === currentAptName)
            || {{}};
  _currentSido = card.sido || '';
  document.getElementById('modalTitle').textContent = currentAptName;
  document.getElementById('modalSub').textContent =
    `${{card.sido||''}} ${{card.gugun||''}} ${{card.dong||''}} · 건축 ${{card.build_year||''}}년`;

  /* 탭 항상 1번(추이)으로 초기화 */
  document.querySelectorAll('.main-tab').forEach((b,i) => b.classList.toggle('active', i===0));
  document.querySelectorAll('.tab-panel').forEach((p,i) => p.classList.toggle('active', i===0));
  if (cmpChartInst) {{ cmpChartInst.destroy(); cmpChartInst = null; }}

  /* 평형 타입 목록 (면적 오름차순) */
  const types = [...new Set(all.map(p => p.ak))]
    .sort((a,b) => parseFloat(a) - parseFloat(b));

  /* 탭 렌더 */
  const tabsEl = document.getElementById('typeTabs');
  tabsEl.innerHTML = ['전체', ...types].map((t, i) => {{
    const color = i === 0 ? '#03c75a' : PALETTE[(i-1) % PALETTE.length];
    return `<button class="type-tab${{t==='전체'?' active':''}}"
      data-type="${{t}}"
      style="${{t==='전체' ? 'background:#03c75a;color:white;' : ''}}"
      onclick="switchType('${{t}}', '${{color}}')"
    >${{i===0?'전체':''}}${{i>0?`<span class="dot" style="background:${{color}}"></span>`:''}}${{t==='전체'?'':fmtArea(t)}}</button>`;
  }}).join('');

  /* defaultType이 실제 존재하는 타입인지 확인, 없으면 '전체' */
  const initType = types.includes(defaultType) ? defaultType : '전체';
  currentActiveType = initType;

  /* 탭 버튼 active 상태 반영 */
  document.querySelectorAll('.type-tab').forEach(btn => {{
    const isActive = btn.dataset.type === initType;
    btn.classList.toggle('active', isActive);
    if (isActive) {{
      btn.style.background = initType === '전체' ? '#03c75a' : PALETTE[types.indexOf(initType) % PALETTE.length];
      btn.style.color = 'white';
    }}
  }});

  const initFiltered = initType === '전체' ? all : all.filter(p => p.ak === initType);
  renderChart(all, types, initType);
  renderMeta(initFiltered);
  renderTable(all, initType, types);

  document.getElementById('modalOverlay').classList.add('open');
  document.body.style.overflow = 'hidden';
}}

/* ── 탭 전환 ── */
function switchType(type, color) {{
  currentActiveType = type;
  const allRaw = SPARK_MAP[currentAptName] || [];
  const all = currentGugun ? allRaw.filter(p => p.g === currentGugun) : allRaw;
  const types = [...new Set(all.map(p => p.ak))]
    .sort((a,b) => parseFloat(a) - parseFloat(b));

  document.querySelectorAll('.type-tab').forEach(btn => {{
    const t = btn.dataset.type;
    btn.classList.toggle('active', t === type);
    if (t === type) {{
      btn.style.background = type === '전체' ? '#03c75a' : color;
      btn.style.color = 'white';
    }} else {{
      btn.style.background = '';
      btn.style.color = '';
    }}
  }});

  const filtered = type === '전체' ? all : all.filter(p => p.ak === type);
  renderChart(all, types, type);
  renderMeta(filtered);
  renderTable(all, type, types);
}}

/* ── 월별 평균 집계 헬퍼 ── */
function monthlyAvg(pts) {{
  const m = {{}};
  pts.forEach(p => {{
    const ym = p.d.slice(0, 7); // "2026.04"
    if (!m[ym]) m[ym] = [];
    m[ym].push(p.v);
  }});
  const months = Object.keys(m).sort();
  return {{
    labels: months,
    values: months.map(ym => Math.round(m[ym].reduce((a,b)=>a+b,0)/m[ym].length)),
  }};
}}

/* ── 차트 렌더 (SPARK_MAP 포맷: {{d,v,ak}}) ── */
function renderChart(all, types, activeType) {{
  if (chartInst) chartInst.destroy();

  let datasets, labels;

  if (activeType === '전체') {{
    /* 전체: 평형별 멀티 라인 (월평균) */
    const allMonths = [...new Set(all.map(p => p.d.slice(0,7)))].sort();
    labels = allMonths;
    datasets = types.map((type, i) => {{
      const color = PALETTE[i % PALETTE.length];
      const {{labels: ml, values: mv}} = monthlyAvg(all.filter(p => p.ak === type));
      const monthMap = {{}};
      ml.forEach((m, j) => monthMap[m] = mv[j]);
      return {{
        label: fmtArea(type),
        data: allMonths.map(m => monthMap[m] ?? null),
        borderColor: color,
        backgroundColor: color + '18',
        borderWidth: 2,
        pointRadius: 2,
        pointBackgroundColor: color,
        tension: 0.4,
        spanGaps: true,
      }};
    }});
  }} else {{
    /* 선택 평형: 월평균 부드러운 단일 라인 + 실거래 점 오버레이 */
    const typeData = all.filter(p => p.ak === activeType);
    const {{labels: ml, values: mv}} = monthlyAvg(typeData);
    labels = ml;

    const maxV = Math.max(...mv), minV = Math.min(...mv);
    const lineColor = PALETTE[types.indexOf(activeType) % PALETTE.length] || '#03c75a';

    datasets = [{{
      label: fmtArea(activeType) + ' 월평균',
      data: mv,
      borderColor: lineColor,
      backgroundColor: lineColor + '18',
      borderWidth: 2,
      pointRadius: mv.map(v => (v === maxV || v === minV) ? 6 : 2),
      pointBackgroundColor: mv.map(v => v === maxV ? '#c0392b' : v === minV ? '#888' : lineColor),
      pointBorderColor: '#fff',
      pointBorderWidth: 1.5,
      tension: 0.4,
      fill: true,
      spanGaps: true,
      order: 2,
    }}];

    /* ── 해당 거래 마커 (▼ 역삼각형) ── */
    if (currentTradeDate && currentTradePrice) {{
      const tradeYm = currentTradeDate.slice(0, 7); // "2026.04"
      const tradeData = ml.map(m => m === tradeYm ? currentTradePrice : null);
      const hasMarker = tradeData.some(v => v !== null);
      if (hasMarker) {{
        datasets.push({{
          type: 'line',
          label: '해당 거래',
          data: tradeData,
          showLine: false,
          pointStyle: 'triangle',
          rotation: 180,
          pointRadius: ml.map(m => m === tradeYm ? 13 : 0),
          pointHoverRadius: ml.map(m => m === tradeYm ? 15 : 0),
          pointBackgroundColor: '#ff6b00',
          pointBorderColor: 'white',
          pointBorderWidth: 2,
          spanGaps: false,
          order: 0,
        }});
      }}
    }}
  }}

  const ctx = document.getElementById('priceChart').getContext('2d');
  chartInst = new Chart(ctx, {{
    type: 'line',
    data: {{ labels, datasets }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      interaction: {{ mode: 'index', intersect: false }},
      plugins: {{
        legend: {{
          display: activeType === '전체' && types.length > 1,
          position: 'top',
          labels: {{ font: {{ size: 11 }}, boxWidth: 12, padding: 10 }}
        }},
        tooltip: {{
          callbacks: {{
            label: ctx => {{
              if (ctx.parsed.y === null) return null;
              const prefix = ctx.dataset.label === '해당 거래' ? '▼ 해당 거래' : ctx.dataset.label;
              return ` ${{prefix}}: ${{fmtPrice(ctx.parsed.y)}}`;
            }}
          }}
        }}
      }},
      scales: {{
        x: {{
          ticks: {{ font: {{ size: 10 }}, maxRotation: 45, maxTicksLimit: 24 }},
          grid: {{ color: '#f0f0f0' }}
        }},
        y: {{
          ticks: {{ font: {{ size: 11 }}, callback: v => fmtPrice(v) }},
          grid: {{ color: '#f0f0f0' }}
        }}
      }}
    }}
  }});
}}

/* ── 통계 렌더 (SPARK_MAP 포맷: {{d,v}}) ── */
function renderMeta(trades) {{
  const prices = trades.map(t => t.v).filter(v => v > 0);
  if (!prices.length) {{ document.getElementById('chartMeta').innerHTML = ''; return; }}
  const maxP = Math.max(...prices), minP = Math.min(...prices);
  const avgP = Math.round(prices.reduce((a,b)=>a+b,0)/prices.length);
  const sorted = [...trades].sort((a,b)=>a.d.localeCompare(b.d));
  const diff = sorted.length > 1 ? sorted[sorted.length-1].v - sorted[0].v : 0;
  document.getElementById('chartMeta').innerHTML = `
    <div class="meta-box">
      <div class="meta-label">최고가</div>
      <div class="meta-value up">${{fmtPrice(maxP)}}</div>
    </div>
    <div class="meta-box">
      <div class="meta-label">최저가</div>
      <div class="meta-value down">${{fmtPrice(minP)}}</div>
    </div>
    <div class="meta-box">
      <div class="meta-label">평균 / 변동</div>
      <div class="meta-value">${{fmtPrice(avgP)}}</div>
      <div style="font-size:11px;color:${{diff>=0?'#d0021b':'#03c75a'}};margin-top:2px">${{diff>=0?'+':''}}${{fmtPrice(diff)}}</div>
    </div>
  `;
}}

/* ── 테이블 렌더 (SPARK_MAP 포맷: {{d,v,ak,f,dt}}) ── */
function renderTable(all, activeType, types) {{
  const rows = (activeType === '전체' ? all : all.filter(p => p.ak === activeType))
    .slice().sort((a,b) => b.d.localeCompare(a.d));
  const prices = rows.map(r => r.v);
  const maxP = Math.max(...prices), minP = Math.min(...prices);

  document.getElementById('tradeBody').innerHTML = rows.map(t => {{
    const color = PALETTE[types.indexOf(t.ak) % PALETTE.length] || '#888';
    const cls = t.v === maxP ? 'high' : t.v === minP ? 'low' : '';
    return `<tr>
      <td>${{t.d}}</td>
      <td><span class="dot" style="background:${{color}}"></span>${{fmtArea(t.ak)}}</td>
      <td>${{t.f}}층</td>
      <td class="price-cell ${{cls}}">${{fmtPrice(t.v)}}</td>
      <td>${{t.dt}}</td>
    </tr>`;
  }}).join('');
}}

/* ── 메인 탭 전환 ── */
let cmpChartInst = null;
let _currentSido = '';

function switchMainTab(tab, btn) {{
  document.querySelectorAll('.main-tab').forEach(b => b.classList.remove('active'));
  document.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
  btn.classList.add('active');
  document.getElementById('panel-' + tab).classList.add('active');
  if (tab === 'compare') renderCompare();
}}

/* ── 비교 탭 ── */
function renderCompare() {{
  const aptName = currentAptName;
  const selfCard = DATA.find(d => d.apt_name === aptName && d.gugun === currentGugun)
                || DATA.find(d => d.apt_name === aptName);
  if (!selfCard) return;
  const selfArea  = selfCard.area_key;
  const selfGugun = currentGugun || selfCard.gugun;

  // SPARK_MAP에서 apt별 통계 계산 — 구군별로 분리 (동명 아파트 구분)
  // key = "아파트명||구군"
  const sparkStats = {{}};
  Object.entries(SPARK_MAP).forEach(([name, allPts]) => {{
    const gugunSet = [...new Set(allPts.map(p => p.g).filter(Boolean))];
    (gugunSet.length ? gugunSet : ['']).forEach(g => {{
      const pts  = g ? allPts.filter(p => p.g === g) : allPts;
      const vals = pts.map(p => p.v).filter(v => v > 0);
      if (!vals.length) return;
      // 최다 거래 평형 계산
      const areaCnt = {{}};
      pts.forEach(p => {{ areaCnt[p.ak] = (areaCnt[p.ak]||0) + 1; }});
      const mainArea = Object.entries(areaCnt).sort((a,b)=>b[1]-a[1])[0]?.[0] || '';
      const key = `${{name}}||${{g}}`;
      sparkStats[key] = {{
        name, gugun: g, mainArea,
        avg: vals.reduce((a,b)=>a+b,0)/vals.length,
        max: Math.max(...vals),
        min: Math.min(...vals),
        cnt: vals.length,
        pts,  // 차트용 원본 포인트
      }};
    }});
  }});

  /* ── 평형별 stats 재계산 헬퍼 ── */
  function statsForArea(pts, ak) {{
    const filtered = pts.filter(p => p.ak === ak);
    const vals = filtered.map(p => p.v).filter(v => v > 0);
    if (!vals.length) return null;
    return {{
      avg: vals.reduce((a,b)=>a+b,0)/vals.length,
      max: Math.max(...vals),
      min: Math.min(...vals),
      cnt: vals.length,
      pts: filtered,
    }};
  }}

  const selfKey     = `${{aptName}}||${{selfGugun}}`;
  const selfRaw     = sparkStats[selfKey];
  /* 자신의 stats는 selfArea 기준으로 재계산 */
  const selfAreaStats = selfRaw ? statsForArea(selfRaw.pts, selfArea) : null;
  const selfAvg       = selfAreaStats ? selfAreaStats.avg : 0;

  // ① 같은 평형(selfArea) 기준 유사 가격 (±35%) → 최대 4개
  const sameArea = Object.entries(sparkStats)
    .filter(([k]) => k !== selfKey)
    .map(([k, s]) => {{
      const aStats = statsForArea(s.pts, selfArea);
      if (!aStats || aStats.cnt < 3) return null;
      return {{ ...s, key: k, ...aStats, area: selfArea,
                diff: Math.abs(aStats.avg - selfAvg) / (selfAvg || 1), sameArea: true }};
    }})
    .filter(x => x && x.diff < 0.35)
    .sort((a,b) => a.diff - b.diff)
    .slice(0, 4);

  // ② 부족하면 mainArea 기준으로 채움 (±25%)
  const needed = 4 - sameArea.length;
  const usedKeys = new Set(sameArea.map(x => x.key));
  const otherArea = needed > 0
    ? Object.entries(sparkStats)
        .filter(([k, s]) => k !== selfKey && !usedKeys.has(k))
        .map(([k, s]) => {{
          const aStats = statsForArea(s.pts, s.mainArea);
          if (!aStats || aStats.cnt < 3) return null;
          return {{ ...s, key: k, ...aStats, area: s.mainArea,
                    diff: Math.abs(aStats.avg - selfAvg) / (selfAvg || 1), sameArea: false }};
        }})
        .filter(x => x && x.diff < 0.25)
        .sort((a,b) => a.diff - b.diff)
        .slice(0, needed)
    : [];

  const similar = [...sameArea, ...otherArea];
  const selfEntry = {{
    name: aptName, gugun: selfGugun, mainArea: selfArea, area: selfArea,
    ...(selfAreaStats || {{ avg:0, max:0, min:0, cnt:0, pts:[] }}),
  }};
  const targets = [
    {{ ...selfEntry, isSelf: true }},
    ...similar.map(s => ({{ ...s, isSelf: false }}))
  ];
  const colors = ['#d0021b','#2563eb','#f5a623','#8b5cf6','#06b6d4'];

  // 월별 평균 집계 — 각 target의 pts는 이미 해당 평형으로 필터된 상태
  const monthKeys = {{}};
  targets.forEach(t => {{
    (t.pts || []).forEach(p => {{
      const ym   = p.d.slice(0,7);
      const tKey = `${{t.name}}||${{t.gugun}}`;
      if (!monthKeys[ym]) monthKeys[ym] = {{}};
      if (!monthKeys[ym][tKey]) monthKeys[ym][tKey] = [];
      monthKeys[ym][tKey].push(p.v);
    }});
  }});
  const allMonths = Object.keys(monthKeys).sort();

  const datasets = targets.map((t, i) => {{
    const color = colors[i];
    const tKey  = `${{t.name}}||${{t.gugun}}`;
    return {{
      label: t.name,
      data: allMonths.map(ym => {{
        const vs = monthKeys[ym]?.[tKey];
        if (!vs || !vs.length) return null;
        return Math.round(vs.reduce((a,b)=>a+b,0)/vs.length);
      }}),
      borderColor: color,
      backgroundColor: color + (t.isSelf ? '22' : '10'),
      borderWidth: t.isSelf ? 3 : 1.8,
      pointRadius: 1.5,
      pointBackgroundColor: color,
      tension: 0.35,
      spanGaps: true,
      fill: t.isSelf,
      order: t.isSelf ? 0 : i,
    }};
  }});

  if (cmpChartInst) cmpChartInst.destroy();
  cmpChartInst = new Chart(
    document.getElementById('cmpChart').getContext('2d'), {{
      type: 'line',
      data: {{ labels: allMonths, datasets }},
      options: {{
        responsive: true, maintainAspectRatio: false,
        interaction: {{ mode: 'index', intersect: false }},
        plugins: {{
          legend: {{ display: false }},
          tooltip: {{
            callbacks: {{
              label: ctx => ` ${{ctx.dataset.label}}: ${{fmtPrice(ctx.parsed.y)}}`
            }}
          }}
        }},
        scales: {{
          x: {{ ticks: {{ font: {{ size: 10 }}, maxTicksLimit: 18, maxRotation: 45 }}, grid: {{ color: '#f0f0f0' }} }},
          y: {{ ticks: {{ font: {{ size: 11 }}, callback: v => fmtPrice(v) }}, grid: {{ color: '#f0f0f0' }} }}
        }}
      }}
    }}
  );

  // 범례
  document.getElementById('cmpLegend').innerHTML = targets.map((t, i) =>
    `<div class="cmp-legend-item ${{t.isSelf ? 'cmp-self' : ''}}">
      <span class="cmp-dot" style="background:${{colors[i]}}"></span>
      <span>
        ${{t.name}}
        <span style="font-size:11px;color:#aaa;font-weight:400"> · ${{t.gugun}} ${{fmtArea(t.area)}}${{!t.isSelf && !t.sameArea ? ' ⚠️다른평형' : ''}}</span>
        ${{t.isSelf ? '<b style="color:#d0021b;font-size:11px"> (현재)</b>' : ''}}
      </span>
    </div>`
  ).join('');

  // 통계 테이블 (targets에 이미 avg/max/min/cnt 포함)
  document.getElementById('cmpBody').innerHTML = targets.map((s, i) => `
    <tr class="${{s.isSelf ? 'cmp-self-row' : ''}}">
      <td>
        <span class="cmp-dot" style="background:${{colors[i]}}"></span>
        ${{s.name}}<br>
        <span style="font-size:10px;color:#aaa">${{s.gugun}} · ${{fmtArea(s.area)}}</span>
      </td>
      <td><b>${{fmtPrice(Math.round(s.avg))}}</b></td>
      <td style="color:#d0021b">${{fmtPrice(s.max)}}</td>
      <td style="color:#03c75a">${{fmtPrice(s.min)}}</td>
      <td>${{s.cnt}}건</td>
    </tr>
  `).join('');
}}

function closeModal(e) {{
  if (e.target === document.getElementById('modalOverlay')) closeModalDirect();
}}
function closeModalDirect() {{
  document.getElementById('modalOverlay').classList.remove('open');
  document.body.style.overflow = '';
}}
document.addEventListener('keydown', e => {{ if (e.key === 'Escape') closeModalDirect(); }});

/* ── 금액 포맷 ── */
function fmtPriceMini(v) {{
  const abs = Math.abs(v), sign = v < 0 ? '-' : '';
  if (abs >= 10000) return sign + Math.floor(abs/10000) + '억' + (abs%10000 ? ' '+(abs%10000/10000*10).toFixed(0)+'천' : '');
  return sign + (abs/1000).toFixed(0) + '천만';
}}
function fmtPrice(v) {{
  const abs = Math.abs(v);
  const sign = v < 0 ? '-' : '';
  if (abs >= 10000) {{
    const uk = Math.floor(abs / 10000);
    const rem = abs % 10000;
    return sign + uk + '억' + (rem ? ' ' + rem.toLocaleString() + '만' : '');
  }}
  return sign + abs.toLocaleString() + '만원';
}}

/* ── 엑셀(CSV) 다운로드 ── */
function downloadExcel() {{
  const headers = ['시도','구군','법정동','아파트명','전용면적(㎡)','층','건축년도',
                   '계약년','계약월','계약일','거래금액(만원)','거래유형','중개사소재지','수집일시'];
  const rows = _currentList.map(d => {{
    const parts = d.date ? d.date.split('.') : ['','',''];
    return [
      d.sido, d.gugun, d.dong, d.apt_name,
      d.area_f, d.floor, d.build_year,
      parts[0], parts[1], parts[2],
      d.amount_raw, d.deal_type, d.agent_loc, d.collected,
    ].map(v => {{
      const s = String(v ?? '');
      return s.includes(',') || s.includes('"') || s.includes('\\n') ? `"${{s.replace(/"/g,'""')}}"` : s;
    }});
  }});

  const csv = '\\uFEFF' + [headers, ...rows].map(r => r.join(',')).join('\\n');
  const blob = new Blob([csv], {{ type: 'text/csv;charset=utf-8;' }});
  const url  = URL.createObjectURL(blob);
  const a    = document.createElement('a');
  const now  = new Date();
  const ts   = now.getFullYear() + String(now.getMonth()+1).padStart(2,'0') + String(now.getDate()).padStart(2,'0');
  a.href = url; a.download = `아파트실거래_${{ts}}.csv`;
  document.body.appendChild(a); a.click();
  document.body.removeChild(a); URL.revokeObjectURL(url);
}}

/* ── 랜딩 ── */
let _ldDate  = '';
let _ldGugun = '';

function fmtD(dt) {{
  return `${{dt.getFullYear()}}.${{String(dt.getMonth()+1).padStart(2,'0')}}.${{String(dt.getDate()).padStart(2,'0')}}`;
}}
function prevDay(dateStr) {{
  const [y, m, d] = dateStr.split('.').map(Number);
  const dt = new Date(y, m - 1, d);
  dt.setDate(dt.getDate() - 1);
  return fmtD(dt);
}}

function initLanding() {{
  const today    = new Date();
  const todayStr = fmtD(today);
  _ldDate = todayStr;

  const year = today.getFullYear(), month = today.getMonth();
  const dim  = new Date(year, month + 1, 0).getDate();

  let html = '';
  for (let day = 1; day <= dim; day++) {{
    const dt      = new Date(year, month, day);
    const dtStr   = fmtD(dt);
    const isToday  = dtStr === todayStr;
    const isFuture = dt > today;
    const isActive = dtStr === _ldDate;
    const sub      = isToday ? '오늘' : isFuture ? '비활성' : '';
    html += `<button
      class="ld-date-chip${{isToday ? ' today' : ''}}${{isActive ? ' active' : ''}}"
      data-date="${{dtStr}}"
      ${{isFuture ? 'disabled' : `onclick="selectLdDate('${{dtStr}}')"`}}
    >${{month + 1}}월 ${{day}}일<span class="chip-sub">${{sub}}</span></button>`;
  }}
  document.getElementById('ldDateChips').innerHTML = html;

  const guguns = [...new Set(DATA.map(d => d.gugun))].sort();
  let gh = `<button class="ld-gugun-chip active" onclick="selectLdGugun(this,'')">전체</button>`;
  guguns.forEach(g => {{
    gh += `<button class="ld-gugun-chip" onclick="selectLdGugun(this,'${{g}}')">${{g}}</button>`;
  }});
  document.getElementById('ldGugunChips').innerHTML = gh;
  updateLdInfo();
}}

function selectLdDate(dateStr) {{
  _ldDate = dateStr;
  document.querySelectorAll('.ld-date-chip').forEach(b => b.classList.toggle('active', b.dataset.date === dateStr));
  updateLdInfo();
}}

function selectLdGugun(el, g) {{
  _ldGugun = g;
  document.querySelectorAll('.ld-gugun-chip').forEach(b => b.classList.remove('active'));
  el.classList.add('active');
}}

function updateLdInfo() {{
  const contract = prevDay(_ldDate);
  const el = document.getElementById('ldInfo');
  if (el) el.innerHTML = `<b>${{_ldDate}}</b> 선택 → <b>${{contract}}</b> 체결된 실거래가 표시`;
}}

function applyLanding() {{
  const contract = prevDay(_ldDate);
  const iso      = contract.replace(/\./g, '-');
  document.getElementById('dateFrom').value = iso;
  document.getElementById('dateTo').value   = iso;
  _activeGugun = _ldGugun;
  document.querySelectorAll('.gugun-chip').forEach(c => c.classList.toggle('active', c.dataset.g === _ldGugun));
  document.getElementById('landingOverlay').style.display = 'none';
  applyFilter();
}}

buildSelects();
initLanding();
</script>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"HTML 생성 완료: {output_path}")
    print(f"총 카드: {len(cards)}건")


if __name__ == "__main__":
    excel_path = load_latest_excel()
    base_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(base_dir, "아파트실거래_카드뷰.html")
    generate_html(excel_path, output_path)

    import export_history
    export_history.build(excel_path)
