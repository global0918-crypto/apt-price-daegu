#!/usr/bin/env python3
"""대구 아파트 실거래 수집 → data/transactions.json 저장"""
import requests, xml.etree.ElementTree as ET, json, os, sys
from datetime import datetime, timedelta
from collections import Counter
from zoneinfo import ZoneInfo
from target_date import compute_actual_report_date

API_KEY = os.environ.get("API_KEY", "")
if not API_KEY:
    print("오류: API_KEY 환경변수가 설정되지 않았습니다.", file=sys.stderr)
    sys.exit(1)
DEV_URL = "https://apis.data.go.kr/1613000/RTMSDataSvcAptTradeDev/getRTMSDataSvcAptTradeDev"
STD_URL = "https://apis.data.go.kr/1613000/RTMSDataSvcAptTrade/getRTMSDataSvcAptTrade"

BASE   = os.path.dirname(os.path.abspath(__file__))
DATA   = os.path.join(BASE, "data")
MASTER = os.path.join(DATA, "아파트실거래_마스터.xlsx")
OUTPUT = os.path.join(DATA, "transactions.json")

DISTRICTS = [
    ("중구",   "27110"), ("동구",   "27140"), ("서구",   "27170"),
    ("남구",   "27200"), ("북구",   "27230"), ("수성구", "27260"),
    ("달서구", "27290"), ("달성군", "27710"), ("군위군", "27720"),
]
MASTER_HEADERS = [
    "시도", "구군", "법정동", "아파트명", "전용면적(㎡)", "층", "건축년도",
    "계약년", "계약월", "계약일", "거래금액(만원)", "거래유형", "중개사소재지", "수집일시",
]


# ── API 호출 ──────────────────────────────────────────────────────────
def api_call(url, lawd_cd, ym):
    try:
        resp = requests.get(url, params={
            "serviceKey": API_KEY, "LAWD_CD": lawd_cd,
            "DEAL_YMD": ym, "pageNo": 1, "numOfRows": 1000,
        }, timeout=30)
        resp.raise_for_status()
        return resp.text
    except Exception as e:
        print(f"    오류: {e}")
        return None


def parse_items(xml_text):
    try:
        root = ET.fromstring(xml_text)
        err_cd = root.findtext(".//errCd") or ""
        if err_cd and err_cd not in ("00", "0"):
            print(f"    API 오류코드: {err_cd} {root.findtext('.//errMsg') or ''}")
            return []
        return root.findall(".//item")
    except Exception as e:
        print(f"    파싱오류: {e}")
        return []


def v(item, *tags):
    """XML 요소에서 첫 번째 유효한 값 반환"""
    for t in tags:
        el = item.find(t)
        if el is not None and el.text and el.text.strip():
            return el.text.strip()
    return ""


def parse_rgst(raw):
    """신고일 문자열 → YYYY-MM-DD (실패 시 '')"""
    if not raw:
        return ""
    r = raw.replace("-", "").replace(".", "").replace(" ", "")
    if len(r) == 6 and r.isdigit():          # YY.MM.DD → YYMMDD
        return f"20{r[:2]}-{r[2:4]}-{r[4:]}"
    if len(r) == 8 and r.isdigit():          # YYYYMMDD
        return f"{r[:4]}-{r[4:6]}-{r[6:]}"
    if len(raw) == 10 and raw[4] == "-":     # YYYY-MM-DD
        return raw
    return ""


# ── 역대 최고가 ────────────────────────────────────────────────────────
def load_historical_highs():
    """마스터 Excel → {apt_name|면적㎡: 최고가} (신규 수집 전 상태 기준)"""
    highs = {}
    if not os.path.exists(MASTER):
        print("  마스터 없음 → 신고가 비교 불가")
        return highs
    try:
        import openpyxl
        wb = openpyxl.load_workbook(MASTER, read_only=True, data_only=True)
        if "대구" in wb.sheetnames:
            for row in wb["대구"].iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                try:
                    apt   = str(row[3] or "").strip()
                    area  = float(str(row[4] or 0).strip() or 0)
                    price = int(str(row[10] or 0).replace(",", "").strip() or 0)
                    if apt and area and price:
                        k = f"{apt}|{round(area)}㎡"
                        if highs.get(k, 0) < price:
                            highs[k] = price
                except Exception:
                    pass
        wb.close()
        print(f"  역대 최고가 로드: {len(highs)}개 아파트·평형")
    except ImportError:
        print("  openpyxl 없음")
    except Exception as e:
        print(f"  마스터 로드 오류: {e}")
    return highs


# ── 마스터 Excel 갱신 ──────────────────────────────────────────────────
def update_master(new_rows):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        existing, daegu_rows = set(), []

        if os.path.exists(MASTER):
            wb = openpyxl.load_workbook(MASTER)
            ws = wb["대구"] if "대구" in wb.sheetnames else None
            if ws:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    r = list(row)
                    daegu_rows.append(r)
                    k = (str(r[3]), str(r[7]), str(r[8]).zfill(2),
                         str(r[9]).zfill(2), str(r[10]), str(r[5]))
                    existing.add(k)

        added = 0
        for r in new_rows:
            k = (str(r[3]), str(r[7]), str(r[8]).zfill(2),
                 str(r[9]).zfill(2), str(r[10]), str(r[5]))
            if k not in existing:
                daegu_rows.append(r)
                existing.add(k)
                added += 1

        daegu_rows.sort(key=lambda r: (
            str(r[7]), str(r[8]).zfill(2), str(r[9]).zfill(2)))

        wb2 = openpyxl.Workbook()
        ws2 = wb2.create_sheet("대구")
        ws2.append(MASTER_HEADERS)
        for r in daegu_rows:
            ws2.append(r)

        fill = PatternFill("solid", fgColor="1F4E79")
        font = Font(bold=True, color="FFFFFF", size=10)
        for cell in ws2[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")
        ws2.freeze_panes = "A2"

        if "Sheet" in wb2.sheetnames:
            del wb2["Sheet"]

        os.makedirs(DATA, exist_ok=True)
        wb2.save(MASTER)
        print(f"  마스터 갱신: +{added}건 (누적 {len(daegu_rows)}건)")
    except Exception as e:
        print(f"  마스터 갱신 오류: {e}")


# ── 메인 ──────────────────────────────────────────────────────────────
def _tx_key(t):
    """거래 고유 식별키 (중복/신규 판별용)"""
    return (
        t.get("apt_name", ""),
        t.get("deal_date", ""),
        round(float(t.get("area", 0))),
        int(t.get("floor", 0)),
        int(t.get("amount", 0)),
    )


def load_prev_state(path):
    """이전 transactions.json 로드 → (prev_keys, prev_rgst)"""
    prev_keys, prev_rgst = set(), {}
    if not os.path.exists(path):
        return prev_keys, prev_rgst
    try:
        with open(path, encoding="utf-8") as f:
            prev = json.load(f)
        for t in prev.get("transactions", []):
            k = _tx_key(t)
            prev_keys.add(k)
            if t.get("rgst_date"):
                prev_rgst[k] = t["rgst_date"]
        print(f"  이전 데이터: {len(prev_keys)}건 (rgst_date 보유: {len(prev_rgst)}건)")
    except Exception as e:
        print(f"  이전 데이터 로드 실패: {e}")
    return prev_keys, prev_rgst


def main():
    now    = datetime.now(ZoneInfo("Asia/Seoul"))
    today_str = now.strftime("%Y-%m-%d")

    # 최근 6개월 수집 (이전 계약+최근 신고분 포착)
    months = []
    cur = now.replace(day=1)
    for _ in range(6):
        months.append(cur.strftime("%Y%m"))
        cur = (cur - timedelta(days=1)).replace(day=1)

    print(f"=== 대구 실거래 수집 ({now.strftime('%Y-%m-%d %H:%M')}) ===")

    # 이전 스냅샷 로드 (신규 거래 감지용)
    print("\n[0] 이전 스냅샷 로드")
    prev_keys, prev_rgst = load_prev_state(OUTPUT)
    is_first_run = len(prev_keys) == 0

    print("\n[1] 역대 최고가 로드 (신고가 판별용)")
    highs = load_historical_highs()

    print("\n[2] API 수집 (현재월 + 전월)")
    transactions, master_rows, apt_set = [], [], set()

    for name, code in DISTRICTS:
        for ym in months:
            print(f"  {name} {ym}", end=" → ")

            # Dev API 우선, 실패 시 Standard API 폴백
            xml   = api_call(DEV_URL, code, ym)
            items = parse_items(xml) if xml else []
            is_dev = bool(items)

            if not items:
                xml   = api_call(STD_URL, code, ym)
                items = parse_items(xml) if xml else []

            cnt = 0
            cnt_cancel = 0
            for item in items:
                cdeal_type = v(item, "cdealType") or ""
                cdeal_day  = v(item, "cdealDay")  or ""
                cancelled  = bool(cdeal_day) or cdeal_type.upper() == "O"

                dy = v(item, "dealYear")
                dm = v(item, "dealMonth").zfill(2)
                dd = v(item, "dealDay").zfill(2)

                apt  = v(item, "aptNm")
                dong = v(item, "umdNm")
                if not apt:
                    continue
                apt_set.add(apt)

                rgst = parse_rgst(v(item, "rgstDate", "rgstDe"))

                try:    area  = round(float(v(item, "excluUseAr") or 0), 2)
                except: area  = 0.0
                try:    price = int(v(item, "dealAmount").replace(",", "") or 0)
                except: price = 0
                try:    floor = int(v(item, "floor") or 0)
                except: floor = 0
                try:    by    = int(v(item, "buildYear") or 0)
                except: by    = 0

                if price <= 0:
                    continue

                transactions.append({
                    "gugun":      name,
                    "dong":       dong,
                    "apt_name":   apt,
                    "area":       area,
                    "floor":      floor,
                    "build_year": by,
                    "deal_date":  f"{dy}-{dm}-{dd}",
                    "rgst_date":  rgst,
                    "amount":     price,
                    "deal_type":  v(item, "dealingGbn") or "매매",
                    "cdeal_type": cdeal_type,
                    "cdeal_day":  cdeal_day,
                })

                if cancelled:
                    cnt_cancel += 1
                else:
                    # 해제건은 master에서 제외 (historical_highs 왜곡 방지)
                    master_rows.append([
                        "대구광역시", name, dong, apt,
                        v(item, "excluUseAr"), floor, by,
                        dy, v(item, "dealMonth"), dd,
                        v(item, "dealAmount").replace(",", ""),
                        v(item, "dealingGbn") or "매매",
                        v(item, "estateAgentSggNm"),
                        now.strftime("%Y-%m-%d %H:%M"),
                    ])
                    cnt += 1

            tag = "[Dev]" if is_dev else "[Std]"
            cancel_info = f" (해제 {cnt_cancel}건)" if cnt_cancel else ""
            print(f"{cnt}건 {tag}{cancel_info}")

    # ── 신규 거래 감지 & rgst_date 자동 부여 ─────────────────────────────
    # API가 rgstDate를 ~95% 공백으로 반환하므로,
    # 이전 스냅샷에 없는 거래 = "오늘 최초 확인" → rgst_date = 오늘
    new_today = 0
    for t in transactions:
        k = _tx_key(t)
        if t["rgst_date"]:
            pass                        # API에서 받은 실제 신고일 유지
        elif k in prev_rgst:
            t["rgst_date"] = prev_rgst[k]   # 이전 실행에서 부여한 날짜 복원
        elif not is_first_run and k not in prev_keys:
            t["rgst_date"] = today_str       # 오늘 최초 등장 → 오늘 날짜
            new_today += 1

    print(f"\n[신규 거래 감지] {new_today}건 (rgst_date={today_str} 자동 부여)")
    if is_first_run:
        print("  ※ 최초 실행 — 신규 감지 스킵 (다음 실행부터 적용)")

    # 신고일 내림차순 → 계약일 내림차순
    transactions.sort(
        key=lambda x: (x["rgst_date"] or "0000-00-00", x["deal_date"]),
        reverse=True,
    )

    # 신고일 분포 출력
    dist = Counter(t["rgst_date"] or "없음" for t in transactions)
    print("\n[신고일 분포 Top 10]")
    for d, n in dist.most_common(10):
        print(f"  {d}: {n}건")

    # 실제 표시 타겟일 계산 (오늘 포함 최근 5 영업일 소급)
    print("\n[타겟일 계산]")
    actual_report_date = compute_actual_report_date(transactions)

    # JSON 저장
    os.makedirs(DATA, exist_ok=True)
    payload = {
        "generated_at":       now.isoformat(),
        "actualReportDate":   actual_report_date,
        "total":              len(transactions),
        "apt_names":          sorted(apt_set),
        "historical_highs":   highs,
        "transactions":       transactions,
    }
    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))

    kb = os.path.getsize(OUTPUT) / 1024
    print(f"\n[3] JSON 저장: {kb:.0f} KB, {len(transactions)}건, 아파트 {len(apt_set)}개")
    print(f"  actualReportDate: {actual_report_date}")

    # 마스터 Excel 갱신
    print("\n[4] 마스터 Excel 갱신")
    update_master(master_rows)


if __name__ == "__main__":
    main()
