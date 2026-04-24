"""
호갱노노에서 아파트별 리뷰 수 수집
결과: data/hogangnono_reviews.json
"""
import requests
import json
import os
import time
import glob
import openpyxl
from difflib import SequenceMatcher

BASE = "https://hogangnono.com"
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Referer": "https://hogangnono.com/",
}
OUTPUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "hogangnono_reviews.json")
MASTER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "아파트실거래_마스터.xlsx")


def load_apt_list():
    """엑셀에서 (아파트명, 구군, 법정동) 고유 목록 추출"""
    path = MASTER
    if not os.path.exists(path):
        files = sorted(glob.glob(os.path.join(os.path.dirname(MASTER), "아파트실거래_*.xlsx")), reverse=True)
        if not files:
            return []
        path = files[0]

    wb = openpyxl.load_workbook(path)
    seen = set()
    apts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[3]:
                continue
            key = (str(row[3]), str(row[1]), str(row[2]))  # 아파트명, 구군, 법정동
            if key not in seen:
                seen.add(key)
                apts.append({"name": row[3], "gugun": row[1], "dong": row[2], "sido": row[0]})
    return apts


def search_apt(name, gugun="", dong=""):
    """호갱노노 검색 API로 aptHash 획득 (대구광역시 단지만 허용)"""
    query = f"{gugun} {name}" if gugun else name
    url = f"{BASE}/api/v2/searches/new"
    try:
        r = requests.get(url, params={"query": query, "limit": 10}, headers=HEADERS, timeout=10)
        if r.status_code != 200:
            return None
        data = r.json()
        candidates = data.get("data", {}).get("matched", {}).get("apt", {}).get("list", [])
        if not candidates:
            return None

        # 대구광역시 항목만 허용 + 이름·주소 유사도로 최적 매칭
        best, best_score = None, 0
        for c in candidates:
            addr = c.get("address", "")
            if "대구광역시" not in addr and "대구" not in addr:
                continue
            name_sim = SequenceMatcher(None, name, c.get("name", "")).ratio()
            addr_sim = SequenceMatcher(None, f"{gugun} {dong}", addr).ratio()
            score = name_sim * 0.7 + addr_sim * 0.3
            if score > best_score:
                best_score = score
                best = c

        if best and best_score > 0.5:
            return {"hash": best["id"], "name": best["name"], "address": best.get("address", ""), "score": round(best_score, 3)}
    except Exception as e:
        print(f"  [검색오류] {name}: {e}")
    return None


def get_review_count(apt_hash):
    """리뷰 총 개수 조회 (/reviews/summary 는 비로그인 허용)"""
    url = f"{BASE}/api/v2/apts/{apt_hash}/reviews/summary"
    try:
        r = requests.get(url, headers=HEADERS, timeout=10)
        if r.status_code == 200:
            return r.json().get("data", {}).get("total", 0)
    except Exception as e:
        print(f"  [리뷰오류] {apt_hash}: {e}")
    return 0


def main():
    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)

    # 기존 결과 로드 (재실행 시 스킵)
    existing = {}
    if os.path.exists(OUTPUT):
        with open(OUTPUT, "r", encoding="utf-8") as f:
            existing = json.load(f)
        print(f"기존 캐시: {len(existing)}건")

    apts = load_apt_list()
    print(f"수집 대상: {len(apts)}개 아파트\n")

    results = dict(existing)
    new_count = 0

    for i, apt in enumerate(apts, 1):
        name = apt["name"]
        # count > 0이면 이미 성공 수집 → 스킵. hash=None(매칭실패)은 재시도 허용
        if name in existing and existing[name].get("count", 0) > 0:
            continue

        print(f"[{i}/{len(apts)}] {name} ({apt['gugun']} {apt['dong']})", end=" ", flush=True)
        info = search_apt(name, apt["gugun"], apt["dong"])

        if not info:
            print("→ 매칭 실패")
            results[name] = {"count": 0, "hash": None, "matched": None}
            continue

        count = get_review_count(info["hash"])
        results[name] = {
            "count": count,
            "hash": info["hash"],
            "matched": info["name"],
            "address": info["address"],
        }
        print(f"→ {info['name']} | 리뷰 {count}개")
        new_count += 1

        # 중간 저장 (50건마다)
        if new_count % 50 == 0:
            with open(OUTPUT, "w", encoding="utf-8") as f:
                json.dump(results, f, ensure_ascii=False, indent=2)
            print(f"  [중간저장] {len(results)}건")

        time.sleep(0.3)

    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    matched = sum(1 for v in results.values() if v.get("hash"))
    print(f"\n완료: 전체 {len(results)}건 / 매칭 성공 {matched}건 / 저장: {OUTPUT}")


if __name__ == "__main__":
    main()
