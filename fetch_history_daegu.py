"""
대구 7년치 아파트 실거래 데이터 수집 (마스터 파일 재생성용)
결과: data/아파트실거래_마스터.xlsx
"""
import requests
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from dateutil.relativedelta import relativedelta
import os
import time

API_KEY = os.environ.get("API_KEY", "c0fb777201beb56b4c5e333b94b2e0d3771cfd15247049464ed6161eb19ee6b3")
API_URL = "https://apis.data.go.kr/1613000/RTMSDataSvcAptTrade/getRTMSDataSvcAptTrade"
MASTER_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "아파트실거래_마스터.xlsx")

DAEGU_DISTRICTS = [
    ("중구","27110"),("동구","27140"),("서구","27170"),
    ("남구","27200"),("북구","27230"),("수성구","27260"),
    ("달서구","27290"),("달성군","27710"),
]

HEADERS = ["시도","구군","법정동","아파트명","전용면적(㎡)","층","건축년도",
           "계약년","계약월","계약일","거래금액(만원)","거래유형","중개사소재지","수집일시"]


def month_range(years=7):
    now = datetime.now()
    start = now - relativedelta(years=years)
    months = []
    cur = start.replace(day=1)
    while cur <= now:
        months.append(cur.strftime("%Y%m"))
        cur += relativedelta(months=1)
    return months


def fetch_all_pages(lawd_cd, deal_ymd):
    items = []
    page = 1
    while True:
        params = {
            "serviceKey": API_KEY,
            "LAWD_CD": lawd_cd,
            "DEAL_YMD": deal_ymd,
            "pageNo": page,
            "numOfRows": 1000,
        }
        try:
            resp = requests.get(API_URL, params=params, timeout=30)
            resp.raise_for_status()
            root = ET.fromstring(resp.text)
            page_items = root.findall(".//item")
            items.extend(page_items)
            total = int(root.findtext(".//totalCount") or 0)
            if page * 1000 >= total:
                break
            page += 1
            time.sleep(0.1)
        except Exception as e:
            print(f"    [오류] {lawd_cd} {deal_ymd} p{page}: {e}")
            break
    return items


def get_text(item, tag):
    el = item.find(tag)
    return el.text.strip() if el is not None and el.text else ""


def save_master(rows_daegu, path):
    wb = openpyxl.Workbook()
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF", size=10)
    align = Alignment(horizontal="center", vertical="center")

    ws = wb.create_sheet("대구")
    ws.append(HEADERS)
    for row in rows_daegu:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = fill; cell.font = font; cell.alignment = align
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(path)


def main():
    os.makedirs(os.path.dirname(MASTER_FILE), exist_ok=True)
    months = month_range(7)
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    print(f"=== 대구 7년치 수집 시작 ({months[0]} ~ {months[-1]}) ===")
    print(f"총 {len(months)}개월 × 대구 8구\n")

    rows = []
    seen = set()

    for district_name, lawd_cd in DAEGU_DISTRICTS:
        print(f"  {district_name}", end="", flush=True)
        dist_new = 0
        for ym in months:
            items = fetch_all_pages(lawd_cd, ym)
            for item in items:
                apt   = get_text(item, "aptNm")
                yr    = get_text(item, "dealYear")
                mo    = get_text(item, "dealMonth")
                day   = get_text(item, "dealDay")
                amt   = get_text(item, "dealAmount").replace(",", "")
                floor = get_text(item, "floor")
                key   = (apt, yr, mo, day, amt, floor)
                if key in seen:
                    continue
                seen.add(key)
                rows.append([
                    "대구", district_name,
                    get_text(item, "umdNm"), apt,
                    get_text(item, "excluUseAr"), floor,
                    get_text(item, "buildYear"),
                    yr, mo, day, amt,
                    get_text(item, "dealingGbn"),
                    get_text(item, "estateAgentSggNm"),
                    now_str,
                ])
                dist_new += 1
            time.sleep(0.15)

        rows.sort(key=lambda r: (str(r[7]), str(r[8]).zfill(2), str(r[9]).zfill(2)))
        save_master(rows, MASTER_FILE)
        print(f" +{dist_new}건 (누적 {len(rows)}건)", flush=True)

    print(f"\n대구 총 수집: {len(rows)}건")
    print(f"저장 완료: {MASTER_FILE}")


if __name__ == "__main__":
    main()
